from data.file_managment import get_planilla_climatologica, convert_month
from dash import Output, Input, State, callback, dcc, html, dash_table, no_update
from dash_iconify import DashIconify
from datetime import date, datetime
from cache import data_cache
import dash_mantine_components as dmc
import pandas as pd
import openpyxl
from io import BytesIO


def generacion_planilla_layout():
    """Layout para generacion de planilla climatologica"""
    content = dmc.MantineProvider([
        dmc.Container([
            dmc.Paper(p="md", shadow="sm", radius="md", withBorder=True, children=[
                dmc.Stack(gap="md", children=[
                    dmc.SimpleGrid(
                        children=[
                            dmc.Select(
                                id="station-selector-planilla",
                                label="Selecciona la estacion",
                                placeholder="Selecciona una estacion",
                                value=None,
                                data=[{"value": station, "label": station}
                                      for station in data_cache.get("LISTA_ESTACIONES", [])],
                                size="lg",
                                w=400,
                                searchable=True
                            ),
                            dmc.DatePickerInput(
                                id="month-year-selector-planilla",
                                label="Selecciona mes y ano",
                                placeholder="Selecciona una fecha",
                                minDate=date(1985, 1, 1),
                                value=None,
                                size="lg",
                                w=400,
                                valueFormat="MMMM YYYY"
                            )
                        ],
                        cols=2,
                        spacing="md"
                    ),
                    dmc.Group(justify="flex-end", children=[
                        dmc.Button(
                            "Generar Planilla",
                            id='generar-planilla-btn',
                            leftSection=DashIconify(icon="mdi:file-document", width=20),
                            size="md",
                            variant="filled"
                        )
                    ]),
                    html.Div(id='loading-status-planilla')
                ])
            ]),
        ],
        strategy="grid",
        fluid=True),
        dmc.Container([
            html.Div(id='planilla-table-container', style={'marginTop': '20px'}),
            html.Div(id='export-button-container', style={'marginTop': '20px'}),
            dcc.Download(id="download-planilla-excel")
        ], fluid=True)
    ])

    return content


def transform_data_to_template(df_raw, station_name, year, month):
    """
    Transform raw data from get_planilla_climatologica to template format

    Args:
        df_raw: DataFrame with raw station data (3 rows per day: 7h, 13h, 19h)
        station_name: Name of the station
        year: Year
        month: Month number (1-12)

    Returns:
        DataFrame formatted according to template
    """
    # Load the template
    template_path = "src/template/Planilla de datos andrea.xlsx"
    df_template = pd.read_excel(template_path, sheet_name='Sheet1', header=None)

    # Fill in header information
    df_template.at[5, 2] = station_name  # Estacion name
    df_template.at[6, 17] = convert_month(month)  # Mes
    df_template.at[7, 17] = year  # Ano

    # Fill in metadata from cache
    metadata_df = data_cache.get('METADATA')
    if metadata_df is not None and station_name in metadata_df.index:
        # Get the row for this station (ESTACION is the index)
        station_metadata = metadata_df.loc[station_name]

        # Fill in geographic coordinates and location data
        # Row 6 (index 5): Latitud (col 7), Departamento (col 12)
        if pd.notna(station_metadata.get('LATITUD')):
            df_template.at[5, 7] = station_metadata['LATITUD']
        if pd.notna(station_metadata.get('DEPARTAMENTO')):
            df_template.at[5, 12] = station_metadata['DEPARTAMENTO']

        # Row 7 (index 6): Longitud (col 7), Provincia (col 12)
        if pd.notna(station_metadata.get('LONGITUD')):
            df_template.at[6, 7] = station_metadata['LONGITUD']
        if pd.notna(station_metadata.get('PROVINCIA')):
            df_template.at[6, 12] = station_metadata['PROVINCIA']

        # Row 8 (index 7): Altitud (col 7), Distrito (col 12)
        if pd.notna(station_metadata.get('ALTITUD')):
            df_template.at[7, 7] = station_metadata['ALTITUD']
        if pd.notna(station_metadata.get('DISTRITO')):
            df_template.at[7, 12] = station_metadata['DISTRITO']

    # Get the number of days in the month
    import calendar
    num_days = calendar.monthrange(year, month)[1]

    # Track which rows in template will have data (for SUMA calculations later)
    data_rows = []

    # Data is organized as 3 rows per day (hours 7, 13, 19)
    for day in range(1, num_days + 1):
        # Calculate row indices in raw data (0-indexed)
        base_idx = (day - 1) * 3

        # Check if we have data for this day
        if base_idx + 2 >= len(df_raw):
            break

        # Get the three readings for this day
        reading_7h = df_raw.iloc[base_idx]
        reading_13h = df_raw.iloc[base_idx + 1]
        reading_19h = df_raw.iloc[base_idx + 2]

        # Row index in template - account for SUMA rows
        # Days 1-10: rows 17-26 (no offset)
        # Row 27: SUMA for days 1-10
        # Days 11-20: rows 28-37 (offset +1)
        # Row 38: SUMA for days 11-20
        # Days 21-31: rows 39-49 (offset +2)
        if day <= 10:
            template_row = 17 + day - 1
        elif day <= 20:
            template_row = 17 + day - 1 + 1  # Skip first SUMA row
        else:  # day 21-31
            template_row = 17 + day - 1 + 2  # Skip two SUMA rows

        data_rows.append(template_row)

        # === Temperature Extremes (columns 1-3) ===
        # Col 1 = Maximo, Col 2 = Minimo (swapped from before)
        # TMAX only appears at 19h reading
        if pd.notna(reading_19h.get('TEMPERATURA MAXIMA DIARIA')):
            df_template.at[template_row, 1] = round(reading_19h['TEMPERATURA MAXIMA DIARIA'], 2)

        # TMIN only appears at 7h reading
        if pd.notna(reading_7h.get('TEMPERATURA MINIMA DIARIA')):
            df_template.at[template_row, 2] = round(reading_7h['TEMPERATURA MINIMA DIARIA'], 2)

        # Media Aritmetica for extremes (col 3) - average of max and min
        if pd.notna(df_template.at[template_row, 1]) and pd.notna(df_template.at[template_row, 2]):
            df_template.at[template_row, 3] = round((df_template.at[template_row, 1] + df_template.at[template_row, 2]) / 2, 2)

        # === Termometro Seco (columns 4-7) ===
        # 7h (col 4), 13h (col 5), 19h (col 6)
        df_template.at[template_row, 4] = round(reading_7h['TEMPERATURA DEL BULBO SECO DIARIO'], 2)
        df_template.at[template_row, 5] = round(reading_13h['TEMPERATURA DEL BULBO SECO DIARIO'], 2)
        df_template.at[template_row, 6] = round(reading_19h['TEMPERATURA DEL BULBO SECO DIARIO'], 2)

        # Media Aritmetica (col 7)
        vals = [df_template.at[template_row, 4], df_template.at[template_row, 5], df_template.at[template_row, 6]]
        vals = [v for v in vals if pd.notna(v)]
        if len(vals) > 0:
            df_template.at[template_row, 7] = round(sum(vals) / len(vals), 2)

        # === Termometro Humedo (columns 8-11) ===
        # 7h (col 8), 13h (col 9), 19h (col 10)
        df_template.at[template_row, 8] = round(reading_7h['TEMPERATURA BULBO HUMEDO DIARIA'], 2)
        df_template.at[template_row, 9] = round(reading_13h['TEMPERATURA BULBO HUMEDO DIARIA'], 2)
        df_template.at[template_row, 10] = round(reading_19h['TEMPERATURA BULBO HUMEDO DIARIA'], 2)

        # Media Aritmetica (col 11)
        vals = [df_template.at[template_row, 8], df_template.at[template_row, 9], df_template.at[template_row, 10]]
        vals = [v for v in vals if pd.notna(v)]
        if len(vals) > 0:
            df_template.at[template_row, 11] = round(sum(vals) / len(vals), 2)

        # === Wind Analysis (columns 12-18) - Paired by time ===
        # 7h: Dirección (col 12), Velocidad (col 13)
        df_template.at[template_row, 12] = reading_7h['DIRECCION VIENTO DIARIA']
        if pd.notna(reading_7h.get('VELOCIDAD DEL VIENTO DIARIO')):
            df_template.at[template_row, 13] = round(float(reading_7h['VELOCIDAD DEL VIENTO DIARIO']), 2)

        # 13h: Dirección (col 14), Velocidad (col 15)
        df_template.at[template_row, 14] = reading_13h['DIRECCION VIENTO DIARIA']
        if pd.notna(reading_13h.get('VELOCIDAD DEL VIENTO DIARIO')):
            df_template.at[template_row, 15] = round(float(reading_13h['VELOCIDAD DEL VIENTO DIARIO']), 2)

        # 19h: Dirección (col 16), Velocidad (col 17)
        df_template.at[template_row, 16] = reading_19h['DIRECCION VIENTO DIARIA']
        if pd.notna(reading_19h.get('VELOCIDAD DEL VIENTO DIARIO')):
            df_template.at[template_row, 17] = round(float(reading_19h['VELOCIDAD DEL VIENTO DIARIO']), 2)

        # Velocidad Media (col 18)
        vals = [df_template.at[template_row, 13], df_template.at[template_row, 15], df_template.at[template_row, 17]]
        vals = [v for v in vals if pd.notna(v)]
        if len(vals) > 0:
            df_template.at[template_row, 18] = round(sum(vals) / len(vals), 2)

        # === Precipitation (columns 19-21) ===
        # 7h (col 19), 19h (col 20) - note: PP appears at 7h and 19h only
        if pd.notna(reading_7h.get('PRECIPITACION')):
            df_template.at[template_row, 19] = round(reading_7h['PRECIPITACION'], 2)
        if pd.notna(reading_19h.get('PRECIPITACION')):
            df_template.at[template_row, 20] = round(reading_19h['PRECIPITACION'], 2)

        # Total (col 21) will be calculated after all days are processed
        # Total = 19h of current day + 7h of next day

        # === Nubosidad (Cloud Analysis) - Reorganized by time ===

        # Cantidad total (Octavos) - 7h, 13h, 19h (cols 22, 23, 24)
        # Sum of bajas, medias, altas cantidad, capped at 8 (oktas scale)
        for idx, reading in enumerate([reading_7h, reading_13h, reading_19h]):
            col = 22 + idx
            cantidad_total = 0
            if pd.notna(reading.get('CANTIDAD DE NUBES BAJAS DIARIAS')):
                cantidad_total += float(reading['CANTIDAD DE NUBES BAJAS DIARIAS'])
            if pd.notna(reading.get('CANTIDAD DE NUBES MEDIAS DIARIAS')):
                cantidad_total += float(reading['CANTIDAD DE NUBES MEDIAS DIARIAS'])
            if pd.notna(reading.get('CANTIDAD DE NUBES ALTAS DIARIAS')):
                cantidad_total += float(reading['CANTIDAD DE NUBES ALTAS DIARIAS'])
            df_template.at[template_row, col] = int(min(cantidad_total, 8))

        # 7h observations (cols 25-31)
        # Bajas: Formas (25), Cantidad (26), Altura (27)
        df_template.at[template_row, 25] = reading_7h['FORMA DE NUBES BAJAS DIARIAS']
        if pd.notna(reading_7h.get('CANTIDAD DE NUBES BAJAS DIARIAS')):
            df_template.at[template_row, 26] = int(reading_7h['CANTIDAD DE NUBES BAJAS DIARIAS'])
        df_template.at[template_row, 27] = reading_7h['ALTURA DE NUBES BAJAS DIARIAS']
        # Medias: Formas (28), Cantidad (29)
        df_template.at[template_row, 28] = reading_7h['FORMA DE NUBES MEDIAS DIARIAS']
        if pd.notna(reading_7h.get('CANTIDAD DE NUBES MEDIAS DIARIAS')):
            df_template.at[template_row, 29] = int(reading_7h['CANTIDAD DE NUBES MEDIAS DIARIAS'])
        # Altas: Formas (30), Cantidad (31)
        df_template.at[template_row, 30] = reading_7h['FORMA DE NUBES ALTAS DIARIAS']
        if pd.notna(reading_7h.get('CANTIDAD DE NUBES ALTAS DIARIAS')):
            df_template.at[template_row, 31] = int(reading_7h['CANTIDAD DE NUBES ALTAS DIARIAS'])

        # 13h observations (cols 32-38)
        # Bajas: Formas (32), Cantidad (33), Altura (34)
        df_template.at[template_row, 32] = reading_13h['FORMA DE NUBES BAJAS DIARIAS']
        if pd.notna(reading_13h.get('CANTIDAD DE NUBES BAJAS DIARIAS')):
            df_template.at[template_row, 33] = int(reading_13h['CANTIDAD DE NUBES BAJAS DIARIAS'])
        df_template.at[template_row, 34] = reading_13h['ALTURA DE NUBES BAJAS DIARIAS']
        # Medias: Formas (35), Cantidad (36)
        df_template.at[template_row, 35] = reading_13h['FORMA DE NUBES MEDIAS DIARIAS']
        if pd.notna(reading_13h.get('CANTIDAD DE NUBES MEDIAS DIARIAS')):
            df_template.at[template_row, 36] = int(reading_13h['CANTIDAD DE NUBES MEDIAS DIARIAS'])
        # Altas: Formas (37), Cantidad (38)
        df_template.at[template_row, 37] = reading_13h['FORMA DE NUBES ALTAS DIARIAS']
        if pd.notna(reading_13h.get('CANTIDAD DE NUBES ALTAS DIARIAS')):
            df_template.at[template_row, 38] = int(reading_13h['CANTIDAD DE NUBES ALTAS DIARIAS'])

        # 19h observations (cols 39-45)
        # Bajas: Formas (39), Cantidad (40), Altura (41)
        df_template.at[template_row, 39] = reading_19h['FORMA DE NUBES BAJAS DIARIAS']
        if pd.notna(reading_19h.get('CANTIDAD DE NUBES BAJAS DIARIAS')):
            df_template.at[template_row, 40] = int(reading_19h['CANTIDAD DE NUBES BAJAS DIARIAS'])
        df_template.at[template_row, 41] = reading_19h['ALTURA DE NUBES BAJAS DIARIAS']
        # Medias: Formas (42), Cantidad (43)
        df_template.at[template_row, 42] = reading_19h['FORMA DE NUBES MEDIAS DIARIAS']
        if pd.notna(reading_19h.get('CANTIDAD DE NUBES MEDIAS DIARIAS')):
            df_template.at[template_row, 43] = int(reading_19h['CANTIDAD DE NUBES MEDIAS DIARIAS'])
        # Altas: Formas (44), Cantidad (45)
        df_template.at[template_row, 44] = reading_19h['FORMA DE NUBES ALTAS DIARIAS']
        if pd.notna(reading_19h.get('CANTIDAD DE NUBES ALTAS DIARIAS')):
            df_template.at[template_row, 45] = int(reading_19h['CANTIDAD DE NUBES ALTAS DIARIAS'])

        # === Visibility (columns 46-48) ===
        # 7h, 13h, 19h
        if pd.notna(reading_7h.get('VISIBILIDAD PREVALECIENTE DIARIA')):
            df_template.at[template_row, 46] = round(float(reading_7h['VISIBILIDAD PREVALECIENTE DIARIA']), 2)
        if pd.notna(reading_13h.get('VISIBILIDAD PREVALECIENTE DIARIA')):
            df_template.at[template_row, 47] = round(float(reading_13h['VISIBILIDAD PREVALECIENTE DIARIA']), 2)
        if pd.notna(reading_19h.get('VISIBILIDAD PREVALECIENTE DIARIA')):
            df_template.at[template_row, 48] = round(float(reading_19h['VISIBILIDAD PREVALECIENTE DIARIA']), 2)

    # Calculate precipitation totals (19h of day N + 7h of day N+1)
    # Process all days except the last one
    for day in range(1, num_days):  # num_days-1 iterations (excludes last day)
        # Calculate template row for current day
        if day <= 10:
            current_row = 17 + day - 1
        elif day <= 20:
            current_row = 17 + day - 1 + 1
        else:
            current_row = 17 + day - 1 + 2

        # Calculate template row for next day
        next_day = day + 1
        if next_day <= 10:
            next_row = 17 + next_day - 1
        elif next_day <= 20:
            next_row = 17 + next_day - 1 + 1
        else:
            next_row = 17 + next_day - 1 + 2

        # Total = 19h of current day (col 20) + 7h of next day (col 19)
        val_19h_current = df_template.at[current_row, 20]
        val_7h_next = df_template.at[next_row, 19]

        vals = [val_19h_current, val_7h_next]
        vals = [v for v in vals if pd.notna(v)]
        if len(vals) > 0:
            df_template.at[current_row, 21] = round(sum(vals), 2)

    # Last day has no total (would need next day's 7h reading)

    # Now add SUMA rows every 10 days
    # Template rows: Day 1 = row 17, Day 10 = row 26, SUMA after day 10 = row 27
    # Day 11 = row 28, Day 20 = row 37, SUMA after day 20 = row 38
    # Days 21-31: rows 39-49, SUMA at row 50 (for remaining days)

    # Calculate SUMA for first 10 days (rows 17-26 data, SUMA at row 27)
    if num_days >= 10:
        suma_row = 27
        start_row = 17
        end_row = 26
        df_template.at[suma_row, 0] = 'Suma'
        calculate_suma(df_template, start_row, end_row, suma_row)

    # Calculate SUMA for days 11-20 (rows 28-37 data, SUMA at row 38)
    if num_days >= 20:
        suma_row = 38
        start_row = 28
        end_row = 37
        df_template.at[suma_row, 0] = 'Suma'
        calculate_suma(df_template, start_row, end_row, suma_row)

    # Calculate SUMA for days 21-31 (rows 39 onwards, SUMA at row 50)
    if num_days > 20:
        suma_row = 50
        start_row = 39
        # Calculate the actual last data row for this month
        last_day_row = 39 + (num_days - 21)  # Days 21-31 offset from row 39
        end_row = last_day_row
        df_template.at[suma_row, 0] = 'Suma'
        calculate_suma(df_template, start_row, end_row, suma_row)

    # Calculate TOTAL and MEDIA rows (rows 51 and 52)
    # Total is sum of all days - need to account for SUMA rows when calculating end_row
    total_row = 51
    media_row = 52
    df_template.at[total_row, 0] = 'Total'
    df_template.at[media_row, 0] = 'Media'

    # Calculate the end row for the last day (accounting for SUMA rows)
    if num_days <= 10:
        end_row = 17 + num_days - 1
    elif num_days <= 20:
        end_row = 17 + num_days - 1 + 1
    else:  # num_days > 20
        end_row = 17 + num_days - 1 + 2

    # For TOTAL, we need to sum only the daily data rows, skipping SUMA rows
    # Use the data_rows list we collected earlier
    calculate_total_and_media(df_template, data_rows, total_row, media_row)

    return df_template


def calculate_suma(df_template, start_row, end_row, suma_row):
    """Calculate SUMA for a range of rows"""
    # Columns to exclude from SUMA calculations (only string columns and altura):
    # - Wind directions (strings): 12, 14, 16
    # - Cloud forms (strings): 25, 28, 30, 32, 35, 37, 39, 42, 44
    # - Cloud altura: 27, 34, 41
    #
    # Columns to INCLUDE:
    # - Temperatures: 1-11
    # - Wind velocities: 13, 15, 17, 18
    # - Precipitation: 19-21
    # - Cantidad total (Octavos): 22, 23, 24
    # - Cloud cantidad: 26, 29, 31, 33, 36, 38, 40, 43, 45
    # - Visibility: 46, 47, 48
    excluded_cols = set([12, 14, 16])  # Wind directions (strings)
    excluded_cols.update([25, 28, 30, 32, 35, 37, 39, 42, 44])  # Cloud forms (strings)
    excluded_cols.update([27, 34, 41])  # Cloud altura

    for col in range(1, 49):  # Columns 1-48
        if col in excluded_cols:
            # Keep NaN for excluded columns in SUMA rows
            df_template.at[suma_row, col] = None
            continue

        try:
            values = []
            for r in range(start_row, end_row + 1):
                val = df_template.at[r, col]
                if pd.notna(val) and isinstance(val, (int, float)):
                    values.append(val)

            if len(values) > 0:
                df_template.at[suma_row, col] = round(sum(values), 2)
        except:
            pass


def calculate_total_and_media(df_template, data_rows, total_row, media_row):
    """Calculate TOTAL and MEDIA for all days (using list of row indices)"""
    # Columns to exclude from TOTAL and MEDIA (only string columns and altura):
    # - Wind directions (strings): 12, 14, 16
    # - Cloud forms (strings): 25, 28, 30, 32, 35, 37, 39, 42, 44
    # - Cloud altura: 27, 34, 41
    #
    # Columns to INCLUDE:
    # - Temperatures: 1-11
    # - Wind velocities: 13, 15, 17, 18
    # - Precipitation: 19-21
    # - Cantidad total (Octavos): 22, 23, 24
    # - Cloud cantidad: 26, 29, 31, 33, 36, 38, 40, 43, 45
    # - Visibility: 46, 47, 48
    excluded_cols = set([12, 14, 16])  # Wind directions (strings)
    excluded_cols.update([25, 28, 30, 32, 35, 37, 39, 42, 44])  # Cloud forms (strings)
    excluded_cols.update([27, 34, 41])  # Cloud altura

    for col in range(1, 49):  # Columns 1-48
        if col in excluded_cols:
            # Keep NaN for excluded columns in TOTAL and MEDIA rows
            df_template.at[total_row, col] = None
            df_template.at[media_row, col] = None
            continue

        try:
            values = []
            # Iterate only over the data rows (skips SUMA rows)
            for r in data_rows:
                val = df_template.at[r, col]
                if pd.notna(val) and isinstance(val, (int, float)):
                    values.append(val)

            if len(values) > 0:
                df_template.at[total_row, col] = round(sum(values), 2)
                df_template.at[media_row, col] = round(sum(values) / len(values), 2)
        except:
            pass


@callback(
    [Output('planilla-table-container', 'children'),
     Output('export-button-container', 'children'),
     Output('loading-status-planilla', 'children')],
    Input('generar-planilla-btn', 'n_clicks'),
    [State('station-selector-planilla', 'value'),
     State('month-year-selector-planilla', 'value')],
    prevent_initial_call=True
)
def generate_planilla(n_clicks, station, fecha):
    """Generate the climatological form with data"""
    if not n_clicks or not station or not fecha:
        return no_update, no_update, dmc.Alert(
            "Por favor selecciona una estacion y una fecha",
            color="red",
            icon=DashIconify(icon="mdi:alert-circle")
        )

    try:
        # Parse date
        fecha_obj = datetime.fromisoformat(fecha)
        year = fecha_obj.year
        month = fecha_obj.month

        # Fetch raw data
        df_raw = get_planilla_climatologica(station, year, month)

        # Transform to template format
        df_filled = transform_data_to_template(df_raw, station, year, month)

        # Convert to display format
        # Replace NaN with empty strings for display
        df_display = df_filled.fillna('')

        # Create DataTable
        table = dash_table.DataTable(
            data=df_display.to_dict('records'),
            columns=[{"name": str(i), "id": str(i)} for i in df_display.columns],
            style_table={'overflowX': 'auto'},
            style_cell={
                'textAlign': 'left',
                'padding': '5px',
                'minWidth': '50px',
                'maxWidth': '180px',
                'whiteSpace': 'normal',
                'height': 'auto',
            },
            style_header={
                'backgroundColor': 'rgb(230, 230, 230)',
                'fontWeight': 'bold'
            },
            style_data_conditional=[
                {
                    'if': {'row_index': 'odd'},
                    'backgroundColor': 'rgb(248, 248, 248)'
                }
            ],
            id='planilla-datatable'
        )

        # Export button
        export_btn = dmc.Group([
            dmc.Button(
                "Exportar a Excel",
                id='export-excel-btn',
                leftSection=DashIconify(icon="mdi:file-excel", width=20),
                size="md",
                variant="filled",
                color="green"
            )
        ])

        # Store the filled dataframe in a hidden div for export
        storage = html.Div(
            id='planilla-data-storage',
            style={'display': 'none'},
            **{'data-station': station, 'data-year': str(year), 'data-month': str(month)}
        )

        success_msg = dmc.Alert(
            f"Planilla generada exitosamente para {station} - {convert_month(month)} {year}",
            color="green",
            icon=DashIconify(icon="mdi:check-circle")
        )

        return [table, storage], export_btn, success_msg

    except Exception as e:
        error_msg = dmc.Alert(
            f"Error al generar planilla: {str(e)}",
            color="red",
            icon=DashIconify(icon="mdi:alert-circle")
        )
        return no_update, no_update, error_msg


@callback(
    Output('download-planilla-excel', 'data'),
    Input('export-excel-btn', 'n_clicks'),
    [State('station-selector-planilla', 'value'),
     State('month-year-selector-planilla', 'value')],
    prevent_initial_call=True
)
def export_to_excel(n_clicks, station, fecha):
    """Export the filled template to Excel file with preserved formatting"""
    if not n_clicks or not station or not fecha:
        return no_update

    try:
        # Parse date
        fecha_obj = datetime.fromisoformat(fecha)
        year = fecha_obj.year
        month = fecha_obj.month

        # Fetch raw data
        df_raw = get_planilla_climatologica(station, year, month)

        # Transform to template format
        df_filled = transform_data_to_template(df_raw, station, year, month)

        # Load the original template workbook to preserve formatting and merged cells
        template_path = "src/template/Planilla de datos andrea.xlsx"
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active

        # Get all merged cell ranges to avoid writing to them
        merged_ranges = ws.merged_cells.ranges.copy()

        # Update cells with filled data
        # df_filled uses 0-indexed rows and columns
        # openpyxl uses 1-indexed rows and columns
        for row_idx in range(len(df_filled)):
            for col_idx in range(len(df_filled.columns)):
                value = df_filled.iloc[row_idx, col_idx]
                # Only update if value is not NaN
                if pd.notna(value):
                    # Convert to 1-indexed for openpyxl
                    cell_row = row_idx + 1
                    cell_col = col_idx + 1

                    # Check if this cell is part of a merged range
                    cell = ws.cell(row=cell_row, column=cell_col)

                    # Skip if this is a merged cell (not the top-left cell)
                    is_merged_non_top = False
                    for merged_range in merged_ranges:
                        if cell.coordinate in merged_range:
                            # Only write to top-left cell of merged range
                            if cell.coordinate != merged_range.start_cell.coordinate:
                                is_merged_non_top = True
                                break

                    if not is_merged_non_top:
                        ws.cell(row=cell_row, column=cell_col, value=value)

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Return download data
        filename = f"Planilla_{station}_{convert_month(month)}_{year}.xlsx"
        return dcc.send_bytes(output.getvalue(), filename)

    except Exception as e:
        print(f"Error exporting to Excel: {e}")
        import traceback
        traceback.print_exc()
        return no_update
